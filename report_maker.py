import os
import json
import requests
import gspread
from oauth2client.service_account import ServiceAccountCredentials
from openpyxl import Workbook
from openpyxl.styles import Alignment, PatternFill, Font, Border, Side

# 💡 1단계 config 파일에서 설정값들을 가져옵니다.
from config import COLUMNS, WEBHOOK_URL, GCP_SERVICE_ACCOUNT_JSON, GOOGLE_SHEET_URL, TARGET_DATE

# 🌟 Track 코드 → 한글 병기 (구글 시트 표기용. SQLite엔 순수 코드 유지)
from hrdk_law_core.certs import label_track1_type, label_track1_risk, label_track2_code

# 시트에 병기로 표기할 Track 칸 이름
_TRACK_LABELERS = {
    "Track1_취급유형": label_track1_type,
    "Track1_위험도": label_track1_risk,
    "Track2_효용코드": label_track2_code,
}

def _row_for_sheet(info, columns):
    """COLUMNS 순서대로 행을 만들되, Track 칸은 한글 병기로 변환해 시트에 표기.
    (info 원본은 안 건드림 → SQLite 저장은 순수 코드 유지)"""
    row = []
    for c in columns:
        val = info.get(c, "")
        labeler = _TRACK_LABELERS.get(c)
        row.append(labeler(val) if labeler else val)
    return row

# 🌟 [신설 헬퍼] 숫자를 엑셀 열 문자(1->A, 17->Q)로 변환해주는 함수
def get_column_letter(n):
    string = ""
    while n > 0:
        n, remainder = divmod(n - 1, 26)
        string = chr(65 + remainder) + string
    return string

# ==========================================
# 1. 구글 시트 마스터 DB 적재 (통합 Upsert 엔진)
# ==========================================
# 🌟 [고도화] 관제용 상태(status)와 로그(log) 파라미터 기본값 추가
def upload_to_google_sheet(total_len, target_laws, target_date=TARGET_DATE, status="🟢 정상 작동", log="특이사항 없음"):
    """[HRDK LAW-RADAR 오버홀] 국가기술자격 관련 법령 전체 통합 Upsert 및 총괄현황 관제 모니터링 기록"""
    if not GCP_SERVICE_ACCOUNT_JSON or not GOOGLE_SHEET_URL:
        print("  ⚠️ 구글 시트 설정 정보가 없어 적재를 건너뜁니다.")
        return

    try:
        # 인증 로직 (기존과 동일)
        creds_dict = json.loads(GCP_SERVICE_ACCOUNT_JSON.strip(), strict=False)
        scope = ["https://spreadsheets.google.com/feeds", "https://www.googleapis.com/auth/drive"]
        creds = ServiceAccountCredentials.from_json_keyfile_dict(creds_dict, scope)
        client = gspread.authorize(creds)
        spreadsheet = client.open_by_key(GOOGLE_SHEET_URL)

        # 1) 총괄현황표 로깅 [옵션 B] — '처리 대상 날짜' 행에 시도 이력 누적
        #    상태 칸 예: "04:13🔴 → 08:47🔴 → 12:31🟢" (한 줄로 그날 이력 전체 확인)
        #    ⚠️ 행 식별은 '처리 대상 날짜(target_date)'로 — 실행한 날이 아님.
        #       (6/10을 수동 실행하면 6/10 행에 기록되어야 함. 자동 6/17과 분리)
        try:
            from hrdk_law_core.sheets import upsert_daily_summary_row
            # target_date(YYYYMMDD)를 표시용(YYYY-MM-DD)으로 변환
            if target_date and len(str(target_date)) == 8:
                display_date = f"{target_date[:4]}-{target_date[4:6]}-{target_date[6:]}"
            else:
                # 대상 날짜를 모르는 경우(연결 실패 등)에만 오늘 날짜 사용
                from datetime import datetime, timezone, timedelta
                display_date = datetime.now(timezone(timedelta(hours=9))).strftime("%Y-%m-%d")
            # status 문자열에서 심볼만 추출
            symbol = "🟢"
            for s in ("🔴", "🟡", "🟢"):
                if s in status:
                    symbol = s
                    break
            upsert_daily_summary_row(
                spreadsheet,
                sheet_name="총괄현황표",
                target_date_display=display_date,
                cols_before_status=[display_date, total_len, len(target_laws)],
                status_symbol=symbol,
                log=log,
            )
        except Exception as e:
            print(f"  ⚠️ 총괄현황표 관제 데이터 기록 실패: {e}")

        # 2) 🌟 핵심 엔진: 하나의 시트("국가기술자격 관련법령")에 전부 Upsert
        if target_laws:
            try:
                ws_main = spreadsheet.worksheet("국가기술자격 관련법령") # 시트 이름 변경!
                existing_records = ws_main.get_all_records()
                
                max_id_num = 0
                natural_key_map = {}
                
                for idx, record in enumerate(existing_records):
                    mst_id = record.get("MST_ID", "")
                    if mst_id.startswith("HRDK-L-"):
                        try:
                            num = int(mst_id.split("-")[-1])
                            if num > max_id_num: max_id_num = num
                        except: pass
                    
                    nat_key = f"{record.get('법령명','')}|{record.get('근거 조문','')}"
                    natural_key_map[nat_key] = idx + 2

                new_rows_to_append = []

                # 🌟 [개선] COLUMNS 길이에 맞춰 끝 열 범위 문자 자동 계산 (예: 17개면 Q)
                end_col_letter = get_column_letter(len(COLUMNS))

                for info in target_laws:
                    nat_key = f"{info.get('법령명','')}|{info.get('근거 조문','')}"
                    
                    if nat_key in natural_key_map:
                        # Update
                        row_idx = natural_key_map[nat_key]
                        existing_id = existing_records[row_idx - 2].get("MST_ID", "")
                        info["MST_ID"] = existing_id 
                        row_data = _row_for_sheet(info, COLUMNS)
                        
                        # 🌟 하드코딩(A~O)을 유연한 계산식(A~Q)으로 변경하여 워크넷 데이터 누락 방지!
                        ws_main.update(range_name=f'A{row_idx}:{end_col_letter}{row_idx}', values=[row_data]) 
                        print(f"  🔄 [Update] {existing_id}")
                    else:
                        # Insert
                        max_id_num += 1
                        new_id = f"HRDK-L-{max_id_num:04d}" 
                        info["MST_ID"] = new_id
                        row_data = _row_for_sheet(info, COLUMNS)
                        new_rows_to_append.append(row_data)
                        print(f"  ✨ [Insert] {new_id}")

                if new_rows_to_append:
                    ws_main.append_rows(new_rows_to_append)
                    
            except Exception as e:
                print(f"  ⚠️ 국가기술자격 관련법령 시트 적재 실패: {e}")

        print("  ✅ 구글 시트 통합 마스터 DB 적재 및 Upsert 완료!")

    except Exception as e:
        print(f"  ❌ 구글 시트 연동 중 에러: {e}")

# ==========================================
# 1-B. 보류목록 탭 내보내기 (코드가 채움, 담당자는 보기만)
# ==========================================
# 머리말(헤더): 담당자가 직관적으로 이해 + "검토상태"는 담당자가 직접 표기 가능
HELD_SHEET_NAME = "보류목록"
HELD_HEADERS = ["기록일시", "법령명", "시행일자", "소관부처", "보류사유", "법령링크", "검토상태"]


def export_held_laws_to_sheet(kb):
    """SQLite held_laws를 구글 시트 '보류목록' 탭으로 내보냅니다 (담당자 확인용)."""
    if not GCP_SERVICE_ACCOUNT_JSON or not GOOGLE_SHEET_URL:
        return
    try:
        creds_dict = json.loads(GCP_SERVICE_ACCOUNT_JSON.strip(), strict=False)
        scope = ["https://spreadsheets.google.com/feeds", "https://www.googleapis.com/auth/drive"]
        creds = ServiceAccountCredentials.from_json_keyfile_dict(creds_dict, scope)
        client = gspread.authorize(creds)
        spreadsheet = client.open_by_key(GOOGLE_SHEET_URL)

        # 탭이 없으면 생성하고 헤더 작성
        try:
            ws = spreadsheet.worksheet(HELD_SHEET_NAME)
        except gspread.WorksheetNotFound:
            ws = spreadsheet.add_worksheet(title=HELD_SHEET_NAME, rows=1000, cols=len(HELD_HEADERS))
            ws.append_row(HELD_HEADERS)

        # 미검토 보류건만 가져와 시트에 반영 (전체 덮어쓰기 대신 누적 append)
        held = kb.get_held_laws(only_unreviewed=True, limit=500)
        if not held:
            print("  ℹ️ 새로 보류된 법령 없음")
            return

        # 이미 시트에 있는 (법령명) 중복 방지
        existing = set()
        try:
            for r in ws.get_all_records():
                existing.add(str(r.get("법령명", "")))
        except Exception:
            pass

        new_rows = []
        for h in held:
            if h["law_name"] in existing:
                continue
            new_rows.append([
                h.get("created_at", ""), h["law_name"], h.get("enforce_date", ""),
                h.get("ministry", ""), h.get("hold_reason", ""), h.get("law_link", ""),
                "",  # 검토상태 — 담당자가 직접 기입
            ])
        if new_rows:
            ws.append_rows(new_rows)
            print(f"  📋 [보류목록] {len(new_rows)}건 시트 반영")
    except Exception as e:
        print(f"  ⚠️ 보류목록 시트 내보내기 실패: {e}")


# ==========================================
# 1-C. 별칭사전 탭 읽기 (담당자가 편집, 코드는 읽기만)
# ==========================================
ALIAS_SHEET_NAME = "자격명칭_별칭사전"
# 머리말(헤더): 앞 4개는 코드가 읽는 필수 / 뒤 3개는 담당자 관리용
ALIAS_HEADERS = ["구명칭", "현행명칭_2026", "등급", "직무", "등록일", "등록자", "비고"]


def ensure_alias_sheet_exists():
    """별칭사전 탭이 없으면 헤더 + 예시 행과 함께 생성합니다 (최초 1회용)."""
    if not GCP_SERVICE_ACCOUNT_JSON or not GOOGLE_SHEET_URL:
        return
    try:
        creds_dict = json.loads(GCP_SERVICE_ACCOUNT_JSON.strip(), strict=False)
        scope = ["https://spreadsheets.google.com/feeds", "https://www.googleapis.com/auth/drive"]
        creds = ServiceAccountCredentials.from_json_keyfile_dict(creds_dict, scope)
        client = gspread.authorize(creds)
        spreadsheet = client.open_by_key(GOOGLE_SHEET_URL)
        try:
            spreadsheet.worksheet(ALIAS_SHEET_NAME)
        except gspread.WorksheetNotFound:
            ws = spreadsheet.add_worksheet(title=ALIAS_SHEET_NAME, rows=1000, cols=len(ALIAS_HEADERS))
            ws.append_row(ALIAS_HEADERS)
            # 🌟 예시 행 추가 — 담당자가 입력 형식을 헷갈리지 않도록.
            # 구명칭이 '[예시]'로 시작하면 코드가 데이터로 읽지 않고 건너뜁니다.
            ws.append_row([
                "[예시] 전자계산기조직응용기사", "정보처리기사", "기사", "정보기술",
                "2026-07-01", "홍길동", "← 이 줄은 작성 예시입니다. 지우지 말고 아래에 추가하세요.",
            ])
            print(f"  ✅ '{ALIAS_SHEET_NAME}' 탭 생성 (헤더 + 예시 행 포함)")
    except Exception as e:
        print(f"  ⚠️ 별칭사전 탭 확인 실패: {e}")


def read_alias_overrides_from_sheet():
    """
    담당자가 구글 시트 '자격명칭_별칭사전' 탭에 직접 추가한 별칭을 읽어옵니다.
    코어 cert_aliases.csv(기본)에 더해, 담당자가 운영 중 추가한 매핑을 반영합니다.

    안전장치:
      - 구명칭이 '[예시]'로 시작하는 행은 작성 예시이므로 건너뜀
      - 구명칭/현행명이 비어 있으면 건너뜀
      - 구명칭과 현행명이 같으면 (의미 없으므로) 건너뜀
      - 같은 구명칭이 여러 번 나오면 마지막 값 우선 (나중 입력 우선)
    반환: {구명칭: 현행명칭} dict (없으면 빈 dict)
    """
    if not GCP_SERVICE_ACCOUNT_JSON or not GOOGLE_SHEET_URL:
        return {}
    try:
        creds_dict = json.loads(GCP_SERVICE_ACCOUNT_JSON.strip(), strict=False)
        scope = ["https://spreadsheets.google.com/feeds", "https://www.googleapis.com/auth/drive"]
        creds = ServiceAccountCredentials.from_json_keyfile_dict(creds_dict, scope)
        client = gspread.authorize(creds)
        spreadsheet = client.open_by_key(GOOGLE_SHEET_URL)
        ws = spreadsheet.worksheet(ALIAS_SHEET_NAME)
        overrides = {}
        skipped_example = 0
        for r in ws.get_all_records():   # 1행(헤더)은 자동 제외됨
            old = str(r.get("구명칭", "")).strip()
            new = str(r.get("현행명칭_2026", "")).strip()
            # 예시 행 건너뛰기
            if old.startswith("[예시]"):
                skipped_example += 1
                continue
            # 빈칸 / 동일 건너뛰기
            if not old or not new or old == new:
                continue
            overrides[old] = new   # 같은 구명칭이 또 나오면 마지막 값으로 덮임 (나중 우선)
        if skipped_example:
            print(f"    (예시 행 {skipped_example}개 건너뜀)")
        return overrides
    except gspread.WorksheetNotFound:
        return {}
    except Exception as e:
        print(f"  ⚠️ 별칭사전 시트 읽기 실패: {e}")
        return {}


# ==========================================
# 1-D. 우대사항 대장 탭 (법령+조문 단위 현황, 방식 B)
# ==========================================
LEDGER_SHEET_NAME = "우대사항_대장"
LEDGER_HEADERS = ["법령명", "조문", "우대분류", "해당 자격종목",
                  "Track1_취급유형", "Track1_위험도", "Track2_효용코드",
                  "중처법대상", "상태", "최근변경일", "비고"]


def _open_spreadsheet():
    """구글 시트 스프레드시트 객체 반환 (공통 인증)."""
    if not GCP_SERVICE_ACCOUNT_JSON or not GOOGLE_SHEET_URL:
        return None
    creds_dict = json.loads(GCP_SERVICE_ACCOUNT_JSON.strip(), strict=False)
    scope = ["https://spreadsheets.google.com/feeds", "https://www.googleapis.com/auth/drive"]
    creds = ServiceAccountCredentials.from_json_keyfile_dict(creds_dict, scope)
    client = gspread.authorize(creds)
    return client.open_by_key(GOOGLE_SHEET_URL)


def init_ledger_baseline(kb, resolve_fn=None):
    """
    [최초 1회] 우대사항 대장 기준선을 깝니다.
    이미 데이터가 있으면 건너뜁니다 (덮어쓰기 방지 = 방식 B).
    """
    try:
        ss = _open_spreadsheet()
        if ss is None:
            return
        try:
            ws = ss.worksheet(LEDGER_SHEET_NAME)
            existing = ws.get_all_values()
            if len(existing) > 1:
                print(f"  ℹ️ 우대사항 대장에 이미 {len(existing)-1}행 존재 → 기준선 적재 생략")
                return
        except gspread.WorksheetNotFound:
            ws = ss.add_worksheet(title=LEDGER_SHEET_NAME, rows=5000, cols=len(LEDGER_HEADERS))
            ws.append_row(LEDGER_HEADERS)

        rows = kb.build_ledger_rows(resolve_fn=resolve_fn)
        # 배치로 한 번에 기록 (속도)
        data = [_row_for_sheet(r, LEDGER_HEADERS) for r in rows]
        if data:
            ws.append_rows(data)
            print(f"  📒 우대사항 대장 기준선 {len(data)}행 적재 완료")
    except Exception as e:
        print(f"  ⚠️ 우대사항 대장 기준선 적재 실패: {e}")


def fill_ledger_hazard_column(kb):
    """
    [일회성] 대장의 기존 행은 그대로 두고, '중처법대상' 칸만 채웁니다.
    (법령명+조문)으로 DB(build_ledger_rows)와 매칭하여 '대상'을 표시.
    기준선이 이미 깔린 뒤 중처법 칸을 추가한 경우 사용. 전체 덮어쓰기 없음.
    반환: 변경된 행 수.
    """
    try:
        ss = _open_spreadsheet()
        if ss is None:
            print("  ⚠️ 스프레드시트 열기 실패")
            return 0
        ws = ss.worksheet(LEDGER_SHEET_NAME)
        records = ws.get_all_values()
        if len(records) <= 1:
            print("  ℹ️ 대장에 데이터가 없습니다.")
            return 0
        header = records[0]
        try:
            law_col = header.index("법령명")
            art_col = header.index("조문")
            hazard_col = header.index("중처법대상")
        except ValueError as e:
            print(f"  ⚠️ 대장 헤더에 필요한 칸이 없습니다: {e}")
            print(f"     (현재 헤더: {header})")
            return 0

        # DB에서 (법령명, 조문) → 중처법대상 매핑 생성
        rows = kb.build_ledger_rows()
        hazard_map = {}
        for r in rows:
            hazard_map[(r["법령명"], r["조문"])] = r.get("중처법대상", "")

        # 시트 각 행을 매칭해 중처법 칸 업데이트 대상 수집
        from gspread.utils import rowcol_to_a1
        updates = []
        filled = 0
        for i, row in enumerate(records[1:], start=2):
            law = row[law_col] if law_col < len(row) else ""
            art = row[art_col] if art_col < len(row) else ""
            cur = row[hazard_col] if hazard_col < len(row) else ""
            want = hazard_map.get((law, art), "")
            if want and want != cur:
                cell = rowcol_to_a1(i, hazard_col + 1)
                updates.append({"range": cell, "values": [[want]]})
                filled += 1

        if updates:
            ws.batch_update(updates)
            print(f"  ✅ 대장 중처법대상 {filled}개 행 채움 완료")
        else:
            print("  ℹ️ 채울 중처법대상 행이 없습니다 (이미 채워졌거나 매칭 없음).")
        return filled
    except Exception as e:
        print(f"  ⚠️ 대장 중처법대상 채우기 실패: {e}")
        return 0


def apply_cert_rename_to_ledger(old_name, new_name):
    """
    [명칭 변경 반영 - 방식 B] 대장에서 옛 종목명이 든 '해당 자격종목' 칸만
    찾아 현행명으로 교체합니다. 전체 덮어쓰기 없음 → 담당자 메모 보존.
    반환: 변경된 행 수.
    """
    try:
        ss = _open_spreadsheet()
        if ss is None:
            return 0
        ws = ss.worksheet(LEDGER_SHEET_NAME)
        records = ws.get_all_values()
        if len(records) <= 1:
            return 0
        header = records[0]
        try:
            cert_col = header.index("해당 자격종목")
            chg_col = header.index("최근변경일")
        except ValueError:
            return 0

        from datetime import datetime, timezone, timedelta
        today = datetime.now(timezone(timedelta(hours=9))).strftime("%Y-%m-%d")
        changed = 0
        updates = []
        for i, row in enumerate(records[1:], start=2):
            if cert_col < len(row) and old_name in row[cert_col]:
                # 쉼표 구분 목록에서 정확히 그 종목만 치환
                certs = [c.strip() for c in row[cert_col].split(",")]
                new_certs = [new_name if c == old_name else c for c in certs]
                # 중복 제거
                seen, dedup = set(), []
                for c in new_certs:
                    if c not in seen:
                        seen.add(c); dedup.append(c)
                new_val = ", ".join(dedup)
                col_letter = get_column_letter(cert_col + 1)
                chg_letter = get_column_letter(chg_col + 1)
                updates.append({"range": f"{col_letter}{i}", "values": [[new_val]]})
                updates.append({"range": f"{chg_letter}{i}", "values": [[today]]})
                changed += 1
        if updates:
            ws.batch_update(updates)
            print(f"  🔤 대장 종목명 변경 반영: {old_name} → {new_name} ({changed}행)")
        return changed
    except gspread.WorksheetNotFound:
        return 0
    except Exception as e:
        print(f"  ⚠️ 대장 명칭 변경 반영 실패: {e}")
        return 0


# ==========================================
# 2. 엑셀 파일 생성 함수 (시트 1개로 단일화)
# ==========================================
def create_excel_report(target_laws, target_date=TARGET_DATE):
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "국가기술자격 관련법령" # 단일 시트
    
    ws1.append(COLUMNS)
    for row_idx, info in enumerate(target_laws, 2):
        ws1.append(_row_for_sheet(info, COLUMNS))
    for col in ws1.columns:
        ws1.column_dimensions[col[0].column_letter].width = 20

    excel_filename = f"HRDK-LAW-RADAR_일일모니터링_{target_date}.xlsx"
    wb.save(excel_filename)
    return excel_filename

# ==========================================
# 3. 메이크닷컴 웹훅 전송 (기존과 동일)
# ==========================================
def send_webhook_with_file(fname, total, high, simple, target_date=TARGET_DATE):
    if not WEBHOOK_URL: return
    # 🌟 [근본 원인 해결!] 메일/웹훅으로 보낼 때도 사람이 읽기 편한 날짜로 변환해서 쏩니다!
    display_date = f"{target_date[:4]}년 {target_date[4:6]}월 {target_date[6:]}일"
    
    # 이제 Make.com은 "20260428"이 아니라 "2026년 04월 28일" 이라는 데이터를 받게 됩니다!
    # 🏷️ system/source: 두 시스템(RADAR/monitor)을 구분하는 식별값 (메일 제목 분기용)
    summary_data = {
        "system": "HRDK LAW-RADAR",
        "source": "radar",
        "subject": f"[LAW-RADAR] {display_date} 자격증 우대사항 분석 (연관 {high}건)",
        "date": display_date, "total": f"{total}건", "high": f"{high}건", "simple": f"{simple}건"
    }
    try:
        if fname and os.path.exists(fname):
            with open(fname, 'rb') as f:
                requests.post(WEBHOOK_URL, data=summary_data, files={'file': (os.path.basename(fname), f, 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')})
        else:
            requests.post(WEBHOOK_URL, data=summary_data)
        print("  ✅ 웹훅 전송 성공!")
    except Exception as e: print(f"  ❌ 웹훅 에러: {e}")
